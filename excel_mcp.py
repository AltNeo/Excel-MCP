import pandas as pd
from fastmcp import FastMCP
from typing import List, Optional, Dict, Any
import json
import os
from openpyxl import load_workbook
import re

mcp = FastMCP(name="ExcelMCP")

@mcp.tool
def read_csv_range(
    file_path: str,
    start_row: Optional[int] = 0,
    end_row: Optional[int] = None,
    columns: Optional[List[str]] = None,
    delimiter: Optional[str] = ',',
    encoding: Optional[str] = 'utf-8',
) -> str:
    """
    Read specific rows/columns from CSV files.
    
    :param file_path: Path to CSV file.
    :param start_row: Starting row (0-based, default: 0).
    :param end_row: Ending row (exclusive, default: all rows).
    :param columns: Specific column names to read.
    :param delimiter: CSV delimiter (default: ',').
    :param encoding: File encoding (default: 'utf-8').
    :return: A JSON string representing the read data.
    """
    try:
        nrows = end_row - start_row if end_row is not None else None
        df = pd.read_csv(
            file_path,
            skiprows=range(1, start_row + 1), # skiprows is 1-based for rows after header
            nrows=nrows,
            usecols=columns,
            sep=delimiter,
            encoding=encoding,
            header=0
        )
        return df.to_json(orient='records')
    except Exception as e:
        return f"Error reading CSV file: {e}"

@mcp.tool
def get_excel_sheet_names(file_path: str) -> str:
    """
    Get all sheet names from an Excel file.
    
    :param file_path: Path to Excel file.
    :return: A JSON string containing the list of sheet names.
    """
    try:
        # Read Excel file to get sheet names
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        return pd.Series(sheet_names).to_json(orient='values')
    except Exception as e:
        return f"Error reading Excel file sheet names: {e}"

@mcp.tool
def get_sheet_info(file_path: str, sheet_name: Optional[str] = 0) -> str:
    """
    Get detailed information about a specific sheet including dimensions, column names, and data types.
    
    :param file_path: Path to Excel file.
    :param sheet_name: Sheet name or index (default: first sheet).
    :return: A JSON string containing sheet metadata.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        
        info = {
            "sheet_name": sheet_name if isinstance(sheet_name, str) else f"Sheet_{sheet_name}",
            "dimensions": {
                "rows": len(df),
                "columns": len(df.columns)
            },
            "columns": {
                "names": df.columns.tolist(),
                "data_types": df.dtypes.astype(str).to_dict()
            },
            "memory_usage": df.memory_usage(deep=True).sum(),
            "has_null_values": df.isnull().any().any(),
            "null_counts": df.isnull().sum().to_dict()
        }
        
        return json.dumps(info, indent=2)
    except Exception as e:
        return f"Error getting sheet info: {e}"

@mcp.tool
def read_excel_range(
    file_path: str,
    sheet_name: Optional[str] = 0,
    start_row: Optional[int] = 0,
    end_row: Optional[int] = None,
    start_col: Optional[int] = 0,
    end_col: Optional[int] = None,
    columns: Optional[List[str]] = None,
    cell_range: Optional[str] = None, # Not yet implemented
) -> str:
    """
    Read specific ranges from Excel files.
    Note: cell_range parameter is not yet implemented.
    
    :param file_path: Path to Excel file.
    :param sheet_name: Sheet name (default: first sheet).
    :param start_row: Starting row (0-based).
    :param end_row: Ending row (exclusive).
    :param start_col: Starting column (0-based).
    :param end_col: Ending column (exclusive).
    :param columns: Specific column names to read.
    :param cell_range: Excel-style range (e.g., "A1:C10").
    :return: A JSON string representing the read data.
    """
    try:
        usecols = None
        if columns:
            usecols = columns
        elif start_col is not None and end_col is not None:
            usecols = list(range(start_col, end_col))

        nrows = end_row - start_row if end_row is not None else None
        
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=0,
            skiprows=start_row,
            nrows=nrows,
            usecols=usecols
        )
        return df.to_json(orient='records')
    except Exception as e:
        return f"Error reading Excel file: {e}"

@mcp.tool
def write_excel_data(
    file_path: str,
    data: str,
    sheet_name: Optional[str] = "Sheet1",
    index: Optional[bool] = False
) -> str:
    """
    Write data to an Excel file.
    
    :param file_path: Path to Excel file to create/overwrite.
    :param data: JSON string containing the data to write.
    :param sheet_name: Name of the sheet to create (default: Sheet1).
    :param index: Whether to write row indices (default: False).
    :return: Success or error message.
    """
    try:
        # Parse JSON data
        data_dict = json.loads(data)
        df = pd.DataFrame(data_dict)
        
        # Write to Excel
        df.to_excel(file_path, sheet_name=sheet_name, index=index, engine='openpyxl')
        
        return f"Successfully wrote {len(df)} rows to {file_path} (sheet: {sheet_name})"
    except Exception as e:
        return f"Error writing Excel file: {e}"

@mcp.tool
def write_csv_data(
    file_path: str,
    data: str,
    delimiter: Optional[str] = ',',
    index: Optional[bool] = False,
    encoding: Optional[str] = 'utf-8'
) -> str:
    """
    Write data to a CSV file.
    
    :param file_path: Path to CSV file to create/overwrite.
    :param data: JSON string containing the data to write.
    :param delimiter: CSV delimiter (default: ',').
    :param index: Whether to write row indices (default: False).
    :param encoding: File encoding (default: 'utf-8').
    :return: Success or error message.
    """
    try:
        # Parse JSON data
        data_dict = json.loads(data)
        df = pd.DataFrame(data_dict)
        
        # Write to CSV
        df.to_csv(file_path, sep=delimiter, index=index, encoding=encoding)
        
        return f"Successfully wrote {len(df)} rows to {file_path}"
    except Exception as e:
        return f"Error writing CSV file: {e}"

@mcp.tool
def search_in_sheet(
    file_path: str,
    search_term: str,
    sheet_name: Optional[str] = 0,
    columns: Optional[List[str]] = None,
    case_sensitive: Optional[bool] = False
) -> str:
    """
    Search for a term within an Excel sheet and return matching rows.
    
    :param file_path: Path to Excel file.
    :param search_term: Text to search for.
    :param sheet_name: Sheet name or index (default: first sheet).
    :param columns: Specific columns to search in (default: all columns).
    :param case_sensitive: Whether search should be case sensitive (default: False).
    :return: JSON string containing matching rows.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        
        # Select columns to search
        search_df = df[columns] if columns else df
        
        # Perform search
        if case_sensitive:
            mask = search_df.astype(str).apply(lambda x: x.str.contains(search_term, na=False)).any(axis=1)
        else:
            mask = search_df.astype(str).apply(lambda x: x.str.contains(search_term, case=False, na=False)).any(axis=1)
        
        results = df[mask]
        
        return json.dumps({
            "search_term": search_term,
            "total_matches": len(results),
            "matching_rows": results.to_dict(orient='records')
        }, indent=2)
        
    except Exception as e:
        return f"Error searching in sheet: {e}"

@mcp.tool
def get_data_summary(
    file_path: str,
    sheet_name: Optional[str] = 0,
    columns: Optional[List[str]] = None
) -> str:
    """
    Get statistical summary of data in an Excel sheet.
    
    :param file_path: Path to Excel file.
    :param sheet_name: Sheet name or index (default: first sheet).
    :param columns: Specific columns to analyze (default: all columns).
    :return: JSON string containing statistical summary.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        
        # Select specific columns if provided
        if columns:
            df = df[columns]
        
        # Get basic statistics
        numeric_summary = df.describe().to_dict() if len(df.select_dtypes(include=['number']).columns) > 0 else {}
        
        # Get additional info
        summary = {
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "column_info": {
                "names": df.columns.tolist(),
                "data_types": df.dtypes.astype(str).to_dict(),
                "non_null_counts": df.count().to_dict(),
                "null_counts": df.isnull().sum().to_dict()
            },
            "numeric_statistics": numeric_summary,
            "memory_usage_mb": round(df.memory_usage(deep=True).sum() / (1024*1024), 2)
        }
        
        # Add categorical column info
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            summary["categorical_info"] = {}
            for col in categorical_cols:
                summary["categorical_info"][col] = {
                    "unique_values": df[col].nunique(),
                    "most_common": df[col].value_counts().head(5).to_dict()
                }
        
        return json.dumps(summary, indent=2)
        
    except Exception as e:
        return f"Error getting data summary: {e}"

@mcp.tool
def convert_excel_to_csv(
    excel_path: str,
    csv_path: str,
    sheet_name: Optional[str] = 0,
    delimiter: Optional[str] = ',',
    encoding: Optional[str] = 'utf-8'
) -> str:
    """
    Convert an Excel sheet to CSV format.
    
    :param excel_path: Path to source Excel file.
    :param csv_path: Path to output CSV file.
    :param sheet_name: Sheet name or index to convert (default: first sheet).
    :param delimiter: CSV delimiter (default: ',').
    :param encoding: Output file encoding (default: 'utf-8').
    :return: Success or error message.
    """
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=0)
        df.to_csv(csv_path, sep=delimiter, index=False, encoding=encoding)
        
        return f"Successfully converted Excel sheet to CSV: {len(df)} rows, {len(df.columns)} columns"
    except Exception as e:
        return f"Error converting Excel to CSV: {e}"

@mcp.tool
def convert_csv_to_excel(
    csv_path: str,
    excel_path: str,
    sheet_name: Optional[str] = "Sheet1",
    delimiter: Optional[str] = ',',
    encoding: Optional[str] = 'utf-8'
) -> str:
    """
    Convert a CSV file to Excel format.
    
    :param csv_path: Path to source CSV file.
    :param excel_path: Path to output Excel file.
    :param sheet_name: Name for the Excel sheet (default: Sheet1).
    :param delimiter: CSV delimiter (default: ',').
    :param encoding: Input file encoding (default: 'utf-8').
    :return: Success or error message.
    """
    try:
        df = pd.read_csv(csv_path, sep=delimiter, encoding=encoding)
        df.to_excel(excel_path, sheet_name=sheet_name, index=False, engine='openpyxl')
        
        return f"Successfully converted CSV to Excel: {len(df)} rows, {len(df.columns)} columns"
    except Exception as e:
        return f"Error converting CSV to Excel: {e}"

@mcp.tool
def get_file_info(file_path: str) -> str:
    """
    Get metadata information about a file.
    
    :param file_path: Path to the file.
    :return: JSON string containing file metadata.
    """
    try:
        if not os.path.exists(file_path):
            return f"File not found: {file_path}"
        
        stat = os.stat(file_path)
        file_ext = os.path.splitext(file_path)[1].lower()
        
        info = {
            "file_path": file_path,
            "file_name": os.path.basename(file_path),
            "file_extension": file_ext,
            "size_bytes": stat.st_size,
            "size_mb": round(stat.st_size / (1024 * 1024), 2),
            "created_time": stat.st_ctime,
            "modified_time": stat.st_mtime,
            "is_excel": file_ext in ['.xlsx', '.xls', '.xlsm'],
            "is_csv": file_ext == '.csv'
        }
        
        # If it's an Excel file, get sheet info
        if info["is_excel"]:
            try:
                excel_file = pd.ExcelFile(file_path)
                info["sheet_count"] = len(excel_file.sheet_names)
                info["sheet_names"] = excel_file.sheet_names
            except Exception:
                info["sheet_count"] = "Unknown"
                info["sheet_names"] = []
        
        return json.dumps(info, indent=2)
        
    except Exception as e:
        return f"Error getting file info: {e}"

@mcp.tool
def validate_data_quality(
    file_path: str,
    sheet_name: Optional[str] = 0,
    columns: Optional[List[str]] = None
) -> str:
    """
    Perform comprehensive data quality validation on an Excel sheet or CSV file.
    
    :param file_path: Path to Excel or CSV file.
    :param sheet_name: Sheet name or index for Excel files (default: first sheet).
    :param columns: Specific columns to validate (default: all columns).
    :return: JSON string containing data quality report.
    """
    try:
        # Determine file type and read data
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        elif file_ext == '.csv':
            df = pd.read_csv(file_path)
        else:
            return f"Unsupported file type: {file_ext}"
        
        # Select specific columns if provided
        if columns:
            df = df[columns]
        
        validation_report = {
            "file_info": {
                "file_path": file_path,
                "shape": {"rows": len(df), "columns": len(df.columns)},
                "columns": df.columns.tolist()
            },
            "data_quality_issues": {},
            "summary": {}
        }
        
        issues = validation_report["data_quality_issues"]
        
        # Check for missing values
        null_counts = df.isnull().sum()
        issues["missing_values"] = {
            "columns_with_nulls": null_counts[null_counts > 0].to_dict(),
            "total_null_values": int(df.isnull().sum().sum()),
            "null_percentage": round((df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100, 2)
        }
        
        # Check for duplicates
        duplicate_rows = df.duplicated().sum()
        issues["duplicates"] = {
            "duplicate_row_count": int(duplicate_rows),
            "duplicate_percentage": round((duplicate_rows / len(df)) * 100, 2) if len(df) > 0 else 0
        }
        
        # Check data types and inconsistencies
        issues["data_types"] = {}
        for col in df.columns:
            col_info = {
                "data_type": str(df[col].dtype),
                "unique_values": int(df[col].nunique()),
                "unique_percentage": round((df[col].nunique() / len(df)) * 100, 2) if len(df) > 0 else 0
            }
            
            # Check for mixed data types in object columns
            if df[col].dtype == 'object':
                sample_types = df[col].dropna().apply(type).value_counts()
                if len(sample_types) > 1:
                    col_info["mixed_types"] = sample_types.to_dict()
            
            # Check for outliers in numeric columns
            if df[col].dtype in ['int64', 'float64']:
                Q1 = df[col].quantile(0.25)
                Q3 = df[col].quantile(0.75)
                IQR = Q3 - Q1
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                outliers = ((df[col] < lower_bound) | (df[col] > upper_bound)).sum()
                col_info["outliers"] = {
                    "count": int(outliers),
                    "percentage": round((outliers / len(df)) * 100, 2) if len(df) > 0 else 0
                }
            
            issues["data_types"][col] = col_info
        
        # Generate summary
        validation_report["summary"] = {
            "overall_quality_score": _calculate_quality_score(issues, len(df)),
            "critical_issues": _identify_critical_issues(issues),
            "recommendations": _generate_recommendations(issues)
        }
        
        return json.dumps(validation_report, indent=2)
        
    except Exception as e:
        return f"Error validating data quality: {e}"

def _calculate_quality_score(issues: Dict, total_rows: int) -> float:
    """Calculate overall data quality score (0-100)."""
    score = 100.0
    
    # Deduct for missing values
    if issues["missing_values"]["null_percentage"] > 0:
        score -= min(issues["missing_values"]["null_percentage"] * 2, 30)
    
    # Deduct for duplicates
    if issues["duplicates"]["duplicate_percentage"] > 0:
        score -= min(issues["duplicates"]["duplicate_percentage"] * 1.5, 20)
    
    # Deduct for data type issues
    for col, col_info in issues["data_types"].items():
        if "mixed_types" in col_info:
            score -= 5
        if "outliers" in col_info and col_info["outliers"]["percentage"] > 10:
            score -= min(col_info["outliers"]["percentage"] * 0.5, 10)
    
    return max(0, round(score, 1))

def _identify_critical_issues(issues: Dict) -> List[str]:
    """Identify critical data quality issues."""
    critical = []
    
    if issues["missing_values"]["null_percentage"] > 20:
        critical.append("High percentage of missing values (>20%)")
    
    if issues["duplicates"]["duplicate_percentage"] > 10:
        critical.append("High percentage of duplicate rows (>10%)")
    
    for col, col_info in issues["data_types"].items():
        if "mixed_types" in col_info:
            critical.append(f"Mixed data types in column '{col}'")
        if "outliers" in col_info and col_info["outliers"]["percentage"] > 15:
            critical.append(f"High percentage of outliers in column '{col}' (>15%)")
    
    return critical

def _generate_recommendations(issues: Dict) -> List[str]:
    """Generate recommendations for data quality improvements."""
    recommendations = []
    
    if issues["missing_values"]["null_percentage"] > 5:
        recommendations.append("Consider handling missing values through imputation or removal")
    
    if issues["duplicates"]["duplicate_row_count"] > 0:
        recommendations.append("Remove duplicate rows to improve data integrity")
    
    for col, col_info in issues["data_types"].items():
        if "mixed_types" in col_info:
            recommendations.append(f"Standardize data types in column '{col}'")
        if "outliers" in col_info and col_info["outliers"]["count"] > 0:
            recommendations.append(f"Investigate outliers in column '{col}' - may indicate data entry errors")
    
    return recommendations

@mcp.tool
def analyze_excel_formulas(
    file_path: str,
    sheet_name: Optional[str] = None,
    max_examples: Optional[int] = 5
) -> str:
    """
    Analyze Excel formulas and show how calculations are derived with examples.
    
    :param file_path: Path to Excel file.
    :param sheet_name: Sheet name to analyze (default: first sheet).
    :param max_examples: Maximum number of example rows to show (default: 5).
    :return: JSON string containing formula analysis with examples.
    """
    try:
        # Load workbook with openpyxl to access formulas
        wb = load_workbook(file_path, data_only=False)
        
        # Get sheet
        if sheet_name is None:
            sheet = wb.active
            sheet_name = sheet.title
        else:
            sheet = wb[sheet_name]
        
        # Also load data values using pandas for comparison
        df_values = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        
        formula_analysis = {
            "file_info": {
                "file_path": file_path,
                "sheet_name": sheet_name,
                "total_rows": sheet.max_row,
                "total_columns": sheet.max_column
            },
            "formula_summary": {
                "total_formulas_found": 0,
                "columns_with_formulas": [],
                "formula_types": {}
            },
            "formula_details": {},
            "examples": []
        }
        
        # Scan for formulas
        formulas_found = {}
        
        # Check header row first
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            headers.append(cell.value if cell.value else f"Column_{col}")
        
        # Scan all cells for formulas
        for row in range(2, min(sheet.max_row + 1, max_examples + 2)):  # Start from row 2 (after header)
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.data_type == 'f' and cell.value:  # Formula cell
                    col_name = headers[col-1] if col-1 < len(headers) else f"Column_{col}"
                    
                    if col_name not in formulas_found:
                        formulas_found[col_name] = {
                            "column_index": col,
                            "formulas": [],
                            "examples": []
                        }
                    
                    # Store formula and example
                    formula_info = {
                        "row": row,
                        "formula": cell.value,
                        "calculated_value": _get_cell_calculated_value(file_path, sheet_name, row, col),
                        "referenced_cells": _extract_cell_references(cell.value)
                    }
                    
                    formulas_found[col_name]["formulas"].append(formula_info)
        
        # Process formula analysis
        formula_analysis["formula_summary"]["total_formulas_found"] = sum(len(col_data["formulas"]) for col_data in formulas_found.values())
        formula_analysis["formula_summary"]["columns_with_formulas"] = list(formulas_found.keys())
        
        # Analyze formula types
        formula_types = {}
        for col_name, col_data in formulas_found.items():
            for formula_info in col_data["formulas"]:
                formula = formula_info["formula"]
                # Extract function names from formulas
                functions = re.findall(r'([A-Z][A-Z0-9]*)\s*\(', formula)
                for func in functions:
                    formula_types[func] = formula_types.get(func, 0) + 1
        
        formula_analysis["formula_summary"]["formula_types"] = formula_types
        
        # Create detailed analysis for each column with formulas
        for col_name, col_data in formulas_found.items():
            col_index = col_data["column_index"]
            
            # Get the most common formula pattern in this column
            formulas = [f["formula"] for f in col_data["formulas"]]
            most_common_formula = max(set(formulas), key=formulas.count) if formulas else None
            
            formula_analysis["formula_details"][col_name] = {
                "column_index": col_index,
                "total_formulas": len(col_data["formulas"]),
                "most_common_formula": most_common_formula,
                "formula_pattern": _generalize_formula_pattern(most_common_formula) if most_common_formula else None,
                "unique_formulas": len(set(formulas))
            }
        
        # Create examples showing derivation
        examples = []
        example_count = 0
        
        for row_idx in range(min(len(df_values), max_examples)):
            pandas_row = df_values.iloc[row_idx]
            excel_row = row_idx + 2  # Excel row (1-indexed, +1 for header)
            
            example = {
                "row_number": excel_row,
                "data": {}
            }
            
            # Show all column values
            for col_idx, col_name in enumerate(headers):
                if col_idx < len(pandas_row):
                    cell_value = pandas_row.iloc[col_idx]
                    
                    # Check if this cell has a formula
                    has_formula = False
                    formula_info = None
                    
                    if col_name in formulas_found:
                        for formula_data in formulas_found[col_name]["formulas"]:
                            if formula_data["row"] == excel_row:
                                has_formula = True
                                formula_info = formula_data
                                break
                    
                    example["data"][col_name] = {
                        "value": cell_value,
                        "has_formula": has_formula,
                        "formula": formula_info["formula"] if formula_info else None,
                        "referenced_cells": formula_info["referenced_cells"] if formula_info else None
                    }
            
            examples.append(example)
            example_count += 1
            
            if example_count >= max_examples:
                break
        
        formula_analysis["examples"] = examples
        
        # Add summary insights
        if formula_analysis["formula_summary"]["total_formulas_found"] == 0:
            formula_analysis["insights"] = ["No formulas found in this sheet. All values appear to be static data."]
        else:
            insights = []
            insights.append(f"Found {formula_analysis['formula_summary']['total_formulas_found']} formulas across {len(formulas_found)} columns")
            
            if formula_types:
                most_used_func = max(formula_types.items(), key=lambda x: x[1])
                insights.append(f"Most commonly used function: {most_used_func[0]} (used {most_used_func[1]} times)")
            
            for col_name, details in formula_analysis["formula_details"].items():
                if details["formula_pattern"]:
                    insights.append(f"Column '{col_name}' uses pattern: {details['formula_pattern']}")
            
            formula_analysis["insights"] = insights
        
        wb.close()
        return json.dumps(formula_analysis, indent=2)
        
    except Exception as e:
        return f"Error analyzing Excel formulas: {e}"

def _get_cell_calculated_value(file_path: str, sheet_name: str, row: int, col: int):
    """Get the calculated value of a cell."""
    try:
        # Load workbook with data_only=True to get calculated values
        wb_values = load_workbook(file_path, data_only=True)
        sheet_values = wb_values[sheet_name]
        value = sheet_values.cell(row=row, column=col).value
        wb_values.close()
        return value
    except Exception:
        return None

def _extract_cell_references(formula: str) -> List[str]:
    """Extract cell references from a formula."""
    if not formula:
        return []
    
    # Pattern to match cell references like A1, B2, $A$1, Sheet1!A1, etc.
    pattern = r'(?:[A-Za-z_][A-Za-z0-9_]*!)?(?:\$?[A-Z]+\$?[0-9]+)'
    references = re.findall(pattern, formula)
    return list(set(references))  # Remove duplicates

def _generalize_formula_pattern(formula: str) -> str:
    """Convert specific formula to a general pattern."""
    if not formula:
        return ""
    
    # Replace specific cell references with patterns
    pattern = re.sub(r'\$?[A-Z]+\$?[0-9]+', '[CELL]', formula)
    pattern = re.sub(r'[A-Za-z_][A-Za-z0-9_]*!', '[SHEET]!', pattern)
    return pattern

if __name__ == "__main__":
    # Run the MCP server using stdio transport for local client communication
    mcp.run(transport="stdio")
